import cv2
import time
import os
from datetime import datetime
from ultralytics import YOLO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- КОНФИГУРАЦИЯ ПРОЕКТА ---
# Путь к весам обученной модели 
MODEL_PATH = "runs/detect/runs/cow_detection/tune_v13/weights/best.pt"

# Источник видео: путь к файлу или индекс камеры (0)
VIDEO_SOURCE = "testvid1.mp4"

# Файл для логирования событий (теперь Excel)
LOG_FILE = "cow_events.xlsx"

# --- НАСТРОЙКИ АЛГОРИТМА ---
# Порог уверенности детектора (0.0 - 1.0).
# 0.3 - оптимально для детекции в сложных условиях.
CONF_THRESH = 0.3

# Интервал записи статистики в Excel (в секундах), чтобы не спамить в файл каждый кадр.
LOG_INTERVAL = 2.0


def init_excel_log(filename):
    """
    Инициализирует Excel файл с заголовками и форматированием.
    """
    try:
        if not os.path.exists(filename):
            # Создаем новый файл с заголовками
            wb = Workbook()
            ws = wb.active
            ws.title = "Мониторинг коров"
            
            # Заголовки столбцов
            headers = ["Timestamp", "Event", "Count", "Message"]
            ws.append(headers)
            
            # Форматирование заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                
                # Автоподбор ширины столбцов
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 20
            
            # Заморозка первой строки
            ws.freeze_panes = "A2"
            
            wb.save(filename)
            print(f"Создан новый Excel файл: {filename}")
        else:
            # Проверяем структуру существующего файла
            wb = load_workbook(filename)
            if "Мониторинг коров" not in wb.sheetnames:
                # Если лист не существует, создаем его
                ws = wb.create_sheet("Мониторинг коров")
                headers = ["Timestamp", "Event", "Count", "Message"]
                ws.append(headers)
            wb.close()
            
    except Exception as e:
        print(f"Предупреждение: Проблема с созданием Excel файла: {e}")


def log_to_excel(filename, event, count, message):
    """
    Добавляет запись в Excel файл.
    """
    try:
        # Загружаем или создаем workbook
        if os.path.exists(filename):
            wb = load_workbook(filename)
        else:
            wb = Workbook()
        
        # Получаем или создаем лист
        if "Мониторинг коров" in wb.sheetnames:
            ws = wb["Мониторинг коров"]
        else:
            ws = wb.create_sheet("Мониторинг коров")
            headers = ["Timestamp", "Event", "Count", "Message"]
            ws.append(headers)
        
        # Добавляем новую строку с данными
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([timestamp, event, count, message])
        
        # Сохраняем файл
        wb.save(filename)
        
    except Exception as e:
        print(f"Ошибка при записи в Excel: {e}")


def setup_excel_logging():
    """
    Настраивает Excel файл для логирования.
    """
    init_excel_log(LOG_FILE)
    
    # Записываем начальное событие
    log_to_excel(LOG_FILE, "START", 0, "Мониторинг запущен")


def run_monitor():
    """
    Основная функция запуска системы мониторинга.
    Инициализирует модель, видеопоток и цикл обработки.
    """
    
    # 1. Проверка окружения
    if not os.path.exists(MODEL_PATH):
        print(f"Error: Файл модели не найден по пути: {MODEL_PATH}")
        print("Убедитесь, что веса скачаны или путь указан верно.")
        return


    print(f"Загрузка модели YOLO: {MODEL_PATH}...")
    try:
        model = YOLO(MODEL_PATH)
    except Exception as e:
        print(f"Critical Error: Не удалось загрузить модель. {e}")
        return

    # 2. Инициализация Excel логирования
    setup_excel_logging()

    # 3. Запуск видеопотока
    cap = cv2.VideoCapture(VIDEO_SOURCE)
    if not cap.isOpened():
        print(f"Error: Не удалось открыть источник видео: {VIDEO_SOURCE}")
        return

    # Настройки визуализации (Шрифты и цвета BGR)
    FONT = cv2.FONT_HERSHEY_SIMPLEX
    COLOR_BOX = (0, 255, 0)      # Зеленый для рамок
    COLOR_TEXT = (255, 255, 255) # Белый для текста интерфейса
    COLOR_BG = (0, 0, 0)         # Черный для подложки

    last_log_time = 0
    last_count = 0

    print("Запуск цикла обработки...")

    # --- ГЛАВНЫЙ ЦИКЛ ---
    while True:
        ret, frame = cap.read()
        if not ret:
            print("Поток завершен (конец видео или ошибка камеры).")
            break

        # 4. Инференс и Трекинг
        # persist=True: критически важно для видео, сохраняет ID объектов между кадрами.
        # verbose=False: отключает мусор в консоли от YOLO.
        results = model.track(frame, persist=True, conf=CONF_THRESH, verbose=False)
        
        # Получаем результаты детекции
        boxes = results[0].boxes
        cow_count = 0

        # 5. Обработка детекций
        # Проверяем, есть ли вообще объекты и присвоены ли им ID
        if boxes.id is not None:
            # Переводим тензоры (GPU) в списки (CPU) для отрисовки
            ids = boxes.id.int().cpu().tolist()
            coords = boxes.xyxy.cpu().tolist()
            
            cow_count = len(ids)

            for b_id, b_xyxy in zip(ids, coords):
                x1, y1, x2, y2 = map(int, b_xyxy)
                
                # Отрисовка рамки объекта
                cv2.rectangle(frame, (x1, y1), (x2, y2), COLOR_BOX, 2)
                
                # Подпись с ID (полезно для отладки трекинга)
                caption = f"#{b_id} Cow"
                
                # Подложка под текст для читаемости
                (w, h), _ = cv2.getTextSize(caption, FONT, 0.6, 1)
                cv2.rectangle(frame, (x1, y1 - 20), (x1 + w, y1), COLOR_BOX, -1)
                cv2.putText(frame, caption, (x1, y1 - 5), FONT, 0.6, (0, 0, 0), 1)

        # 6. GUI Оверлей (Дашборд)
        # Полупрозрачная плашка в левом верхнем углу
        overlay = frame.copy()
        cv2.rectangle(overlay, (0, 0), (300, 120), COLOR_BG, -1)
        frame = cv2.addWeighted(overlay, 0.6, frame, 0.4, 0)
        
        # Вывод статистики на экран
        cv2.putText(frame, f"Total Cows: {cow_count}", (10, 40), FONT, 0.8, COLOR_TEXT, 2)
        cv2.putText(frame, f"Log: {LOG_FILE}", (10, 80), FONT, 0.5, (180, 180, 180), 1)
        cv2.putText(frame, "Press 'q' to quit", (10, 110), FONT, 0.5, (180, 180, 180), 1)

        # 7. Логирование в Excel файл
        # Пишем статус раз в LOG_INTERVAL секунд или при изменении количества
        current_time = time.time()
        if current_time - last_log_time > LOG_INTERVAL or cow_count != last_count:
            log_to_excel(LOG_FILE, "MONITORING", cow_count, f"Количество коров: {cow_count}")
            last_log_time = current_time
            last_count = cow_count

        # Отображение кадра
        cv2.imshow("CowMonitor Pro", frame)

        # Выход по клавише 'q'
        if cv2.waitKey(60) & 0xFF == ord('q'):
            # Логируем остановку
            log_to_excel(LOG_FILE, "STOP", cow_count, "Мониторинг остановлен пользователем")
            break

    # Освобождение ресурсов
    cap.release()
    cv2.destroyAllWindows()


# Точка входа.
# Обязательна для корректной работы multiprocessing в PyTorch на Linux/Windows.
if __name__ == '__main__':
    run_monitor()
