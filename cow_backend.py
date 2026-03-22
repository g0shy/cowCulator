import cv2
import asyncio
import websockets
import base64
import json
import time
import os
from datetime import datetime
from ultralytics import YOLO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- КОНФИГУРАЦИЯ ---
MODEL_PATH = "runs/detect/runs/cow_detection/tune_v13/weights/best.pt"
VIDEO_SOURCE = "testvid1.mp4"  # или 0 для камеры
CONF_THRESH = 0.3
LOG_INTERVAL = 2.0
LOG_FILE = "cow_events.xlsx"

# WebSocket порт
WS_PORT = 8765

# --- Excel функции ---
def init_excel_log(filename):
    try:
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Мониторинг коров"
            headers = ["Timestamp", "Event", "Count", "Message"]
            ws.append(headers)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 20
            ws.freeze_panes = "A2"
            wb.save(filename)
            print(f"Создан новый Excel файл: {filename}")
    except Exception as e:
        print(f"Предупреждение: Проблема с созданием Excel файла: {e}")

def log_to_excel(filename, event, count, message):
    try:
        if os.path.exists(filename):
            wb = load_workbook(filename)
        else:
            wb = Workbook()
        if "Мониторинг коров" in wb.sheetnames:
            ws = wb["Мониторинг коров"]
        else:
            ws = wb.create_sheet("Мониторинг коров")
            headers = ["Timestamp", "Event", "Count", "Message"]
            ws.append(headers)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([timestamp, event, count, message])
        wb.save(filename)
    except Exception as e:
        print(f"Ошибка при записи в Excel: {e}")

# --- WebSocket сервер ---
class CowMonitorServer:
    def __init__(self):
        self.clients = set()
        
    async def register(self, websocket):
        self.clients.add(websocket)
        print(f"Клиент подключен. Всего клиентов: {len(self.clients)}")
        
    async def unregister(self, websocket):
        self.clients.remove(websocket)
        print(f"Клиент отключен. Всего клиентов: {len(self.clients)}")
        
    async def send_frame(self, frame, detections):
        """Отправляет кадр и данные детекции всем клиентам"""
        if not self.clients:
            return
            
        # Кодируем кадр в JPEG, затем в base64
        _, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
        frame_b64 = base64.b64encode(buffer).decode('utf-8')
        
        # Формируем сообщение
        message = {
            'type': 'frame',
            'data': frame_b64,
            'detections': detections
        }
        
        # Отправляем всем клиентам
        for client in self.clients.copy():
            try:
                await client.send(json.dumps(message))
            except:
                await self.unregister(client)
                
    async def send_log_event(self, event, count, message):
        """Отправляет лог-событие клиентам"""
        if not self.clients:
            return
            
        log_msg = {
            'type': 'log',
            'timestamp': datetime.now().strftime("%H:%M:%S"),
            'event': event,
            'count': count,
            'message': message
        }
        
        for client in self.clients.copy():
            try:
                await client.send(json.dumps(log_msg))
            except:
                await self.unregister(client)

async def handle_client(websocket, server):
    """Обработчик WebSocket соединения"""
    await server.register(websocket)
    try:
        async for message in websocket:
            # Обработка команд от клиента
            data = json.loads(message)
            if data.get('command') == 'stop':
                print("Получена команда остановки")
    except:
        pass
    finally:
        await server.unregister(websocket)

async def websocket_server(server):
    """Запуск WebSocket сервера"""
    async def handler(websocket):
        await handle_client(websocket, server)
    
    async with websockets.serve(handler, "localhost", WS_PORT):
        print(f"WebSocket сервер запущен на порту {WS_PORT}")
        await asyncio.Future()

async def process_video(server):
    """Обработка видео и отправка кадров"""
    
    # Инициализация модели
    if not os.path.exists(MODEL_PATH):
        print(f"Error: Файл модели не найден: {MODEL_PATH}")
        return
        
    print(f"Загрузка модели YOLO: {MODEL_PATH}...")
    try:
        model = YOLO(MODEL_PATH)
    except Exception as e:
        print(f"Critical Error: Не удалось загрузить модель. {e}")
        return
        
    # Инициализация Excel
    init_excel_log(LOG_FILE)
    log_to_excel(LOG_FILE, "START", 0, "Мониторинг запущен")
    
    # Открытие видео
    cap = cv2.VideoCapture(VIDEO_SOURCE)
    if not cap.isOpened():
        print(f"Error: Не удалось открыть источник видео: {VIDEO_SOURCE}")
        return
    
    # Настройки отрисовки
    FONT = cv2.FONT_HERSHEY_SIMPLEX
    COLOR_BOX = (0, 255, 0)
    COLOR_TEXT = (255, 255, 255)
    
    last_log_time = 0
    last_count = 0
    
    # Для сглаживания детекций
    last_detections = []
    smoothing_frames = 3  # Количество кадров для сглаживания
    
    print("Запуск цикла обработки...")
    
    # Получаем FPS видео для правильной задержки
    fps = cap.get(cv2.CAP_PROP_FPS)
    if fps <= 0:
        fps = 30
    frame_delay = 1.0 / fps
    
    frame_count = 0
    
    while True:
        ret, frame = cap.read()
        if not ret:
            print("Поток завершен")
            log_to_excel(LOG_FILE, "END", last_count, "Видеопоток завершен")
            await server.send_log_event("END", last_count, "Видеопоток завершен")
            break
        
        frame_count += 1
        
        # Инференс на каждом кадре (убрали frame_skip)
        try:
            results = model.track(frame, persist=True, conf=CONF_THRESH, verbose=False)
        except Exception as e:
            print(f"Ошибка инференса: {e}")
            # Если ошибка, отправляем кадр без детекций
            detection_data = {
                'count': last_count,
                'cows': last_detections,
                'timestamp': datetime.now().strftime("%H:%M:%S")
            }
            await server.send_frame(frame, detection_data)
            await asyncio.sleep(frame_delay)
            continue
        
        # Обработка детекций
        boxes = results[0].boxes
        cow_count = 0
        current_detections = []
        
        if boxes is not None and boxes.id is not None:
            ids = boxes.id.int().cpu().tolist()
            coords = boxes.xyxy.cpu().tolist()
            confs = boxes.conf.cpu().tolist() if boxes.conf is not None else [1.0] * len(ids)
            cow_count = len(ids)
            
            # Отрисовка рамок прямо на кадре
            for b_id, b_xyxy, conf in zip(ids, coords, confs):
                x1, y1, x2, y2 = map(int, b_xyxy)
                
                # Проверяем, что координаты в пределах кадра
                x1 = max(0, x1)
                y1 = max(0, y1)
                x2 = min(frame.shape[1], x2)
                y2 = min(frame.shape[0], y2)
                
                cv2.rectangle(frame, (x1, y1), (x2, y2), COLOR_BOX, 2)
                
                caption = f"#{b_id}"
                (w, h), _ = cv2.getTextSize(caption, FONT, 0.5, 1)
                cv2.rectangle(frame, (x1, y1 - 18), (x1 + w, y1), COLOR_BOX, -1)
                cv2.putText(frame, caption, (x1, y1 - 5), FONT, 0.5, (0, 0, 0), 1)
                
                # Сохраняем данные для фронта
                current_detections.append({
                    'id': b_id,
                    'bbox': [x1, y1, x2, y2],
                    'confidence': round(conf, 2)
                })
        
        # Сглаживание: если детекций нет, но были на прошлых кадрах - показываем последние
        if cow_count == 0 and last_count > 0 and frame_count % smoothing_frames != 0:
            # На некоторых кадрах YOLO может терять объекты, используем последние детекции
            current_detections = last_detections
            cow_count = last_count
        
        # Обновляем последние детекции
        if cow_count > 0:
            last_detections = current_detections
            last_count = cow_count
        
        # Оверлей с инфой на кадре
        cv2.putText(frame, f"Cows: {cow_count}", (10, 30), FONT, 0.7, COLOR_TEXT, 2)
        cv2.putText(frame, f"Frame: {frame_count}", (10, 60), FONT, 0.5, (180, 180, 180), 1)
        
        # Логирование
        current_time = time.time()
        if current_time - last_log_time > LOG_INTERVAL:
            if cow_count != last_count or current_time - last_log_time > LOG_INTERVAL:
                log_to_excel(LOG_FILE, "MONITORING", cow_count, f"Количество коров: {cow_count}")
                await server.send_log_event("MONITORING", cow_count, f"Количество коров: {cow_count}")
                last_log_time = current_time
        
        # Отправляем кадр клиентам
        detection_data = {
            'count': cow_count,
            'cows': current_detections,
            'timestamp': datetime.now().strftime("%H:%M:%S"),
            'frame': frame_count
        }
        await server.send_frame(frame, detection_data)
        
        # Задержка для реального FPS видео
        await asyncio.sleep(frame_delay)
        
    # Освобождение ресурсов
    cap.release()
    print("Мониторинг остановлен")

async def main():
    """Главная асинхронная функция"""
    server = CowMonitorServer()
    
    # Запускаем WebSocket сервер и обработку видео параллельно
    await asyncio.gather(
        websocket_server(server),
        process_video(server)
    )

if __name__ == '__main__':
    # Запускаем асинхронную главную функцию
    asyncio.run(main())